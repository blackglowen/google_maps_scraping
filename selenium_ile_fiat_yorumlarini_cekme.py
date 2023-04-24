import xlsxwriter
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import pyautogui
import time
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.support.wait import WebDriverWait



def google_maps_yorum_bilgileri_cekme(url):
    #url = 'https://www.google.com/maps/place/%C3%96ZKAR+Otomotiv+-+FIAT+Yetkili+Servis/@41.0152584,28.7892399,17z/data=!3m1!4b1!4m6!3m5!1s0x14caa396dc543481:0xc341d989c7a02b2f!8m2!3d41.0152584!4d28.7918202!16s%2Fg%2F11cn5psjms?coh=164777&entry=tt&shorturl=1'
    options = webdriver.ChromeOptions()
    options.add_argument("--user-data-dir=C:\\Users\\AwozeN\\AppData\\Local\\Google\\Chrome\\User Data")  # Path to your chrome profile
    # options.add_argument('--profile-directory=Profile 6') #belirli bir profil olarak belirlemek istersek bu satırı yorum olmaktan çıkarabiliriz. Profil 2 nedir diye merak ediyorsak google chrome profile bilgilerinin bulunduğu dizine gidip(C:\Users\AwozeN\AppData\Local\Google\Chrome\User Data) ordan profile 2 hangi değer karşılık geldiğini bulabiliriz

    chrome_driver_location = Service('C:\\Users\\AwozeN\\PycharmProjects\\bevisible_V1\\chromedriver.exe')  # executable_path çalışmadığı için yani artık kullanılmadığı için hata vermemesi için servis içerisinde kullanıp daha sonra çağırıcaz.

    driver = webdriver.Chrome(service=chrome_driver_location,options=options)  # executable_path değil service= dedik çünkü executable_path artık kullanılmıyor.

    driver.get(url)  # istenilen sayfanın açılmasını/çağırılmasını sağlıyor.




    """Yorumlar butonuna tıklanması için gerekli olan alan bu alan bizi yorumların olduğu sayfaya gönderecektir. Burada direk olarak yorumların link değeri de olabilirdi."""
    driver.find_element(by=By.XPATH,value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[3]/div/div/button[2]').click()

    """ En yeni yorumların çekilebilmesi için tıklanması gereken alan"""
    driver.find_element(by=By.XPATH,value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[3]/div[9]/button').click()

    while True:
        if pyautogui.locateCenterOnScreen("C:\\Users\\AwozeN\\PycharmProjects\\bevisible_V1\\selenium_resimler\\en_yeni.PNG"):
            deger2 = str("C:\\Users\\AwozeN\\PycharmProjects\\bevisible_V1\\selenium_resimler\\en_yeni.PNG")
            pyautogui.click(deger2)
            break


    time.sleep(5) #sayfanın yüklenmesi için biraz bekletiyoruz.
    soup = BeautifulSoup(driver.page_source, "html.parser") # buradaki driver kısmı üst taraf ile farklıdır onu da değiştirmemiz gerekiyor.
    wait = WebDriverWait(driver, 10) #biraz bekletiyoruz sayfanın yüklenmesi için


    num=0
    time.sleep(5)

    """ title kısmının çekildiği yer"""
    title=soup.find('title').text
    # print(soup.prettify())
    while True:
        if num<=10:
            # yorum bilgilerinin çekildiği kısım
            user_name = soup.findAll('div', class_='d4r55')  # kullanıcı ismi
            review_rating = soup.findAll('span', class_='kvMYJc')  # kullanıcının verdiği puan
            # user_review = soup.findAll('span', class_='wiI7pd')  # kullanıcının yaptığı yorum
            user_reviews_create_date = soup.findAll('span', class_='rsqaWe')



            details = []
            user_review_list=[]
            firmanin_cevabi=[]
            firmanin_cevap_verdigi_tarih=[]
            """BU KISIM YORUMLARIN OLMADIĞI DURUMLARDA LİSTE İÇERİSİNE YORUM BULUNMAMAKTADIR ŞEKLİNDE EKLEME YAPMASI İÇİN YAZILMIŞTIR AKSİ HALDE LİSTE İÇERİSİNDEKİ YORUMLAR KAYIYOR YANİ DİĞER KULLANICININ YORUMU BİR ÖNCEKİNE ATANIYOR."""
            yorum_bulma = soup.findAll('div', class_='GHT2ce')
            for rw in range(len(yorum_bulma)):
                cnvrt = str(yorum_bulma[rw])
                print("cnvrt degeri:", cnvrt)
                if "GHT2ce NsCY4" in cnvrt:
                    pass
                else:
                    if "MyEned" in cnvrt:
                        print("yorum vardır.")
                        # yorum içerisinde daha fazla gibi bir button varsa onun tıklanabilmesi için aşağıdaki kod kullanılıyor.
                        if yorum_bulma[rw].find('button', class_='w8nwRe kyuRq'):# eğer ki yorum içerisinde daha fazla gibi yorumun cevabı uzunsa buraya girip tıklanması gerekiyor. daha sonra yorum kısmı çekilmesi lazım.
                            print("daha fazlası bulundu")

                            button = yorum_bulma[rw].find('button', class_='w8nwRe kyuRq')
                            for m in button:
                                print(m.text)
                                if m.text == "Daha fazla":
                                    driver.execute_script('arguments[0].click();', m)
                                    print("daha fazla olan kısım tıklandı")
                            time.sleep(3)
                            user_review_list.append(yorum_bulma[rw].find('span', class_='wiI7pd').text)
                        else:# eğer ki yorum uzun değilse o zaman tıklanacak bir şey yoktur o yüzden direk yorumu çekeriz.
                            user_review_list.append(yorum_bulma[rw].find('span', class_='wiI7pd').text)


                        # user_review_list.append(yorum_bulma[rw].find('span', class_='wiI7pd').text)


                    else:
                        user_review_list.append("yorum bulunmamaktadır.")
                        print("yorum yoktur")
                    if "CDe7pd" in cnvrt:
                        firmanin_cevabi_degeri=yorum_bulma[rw].find('div', class_='CDe7pd')
                        print("firmanın cevabı:",firmanin_cevabi_degeri)
                        firmanin_cevabi.append(firmanin_cevabi_degeri.find('div',class_='wiI7pd').text)
                        firmanin_cevap_verdigi_tarih.append(firmanin_cevabi_degeri.find('span',class_='DZSIDd').text)

                    else:
                        firmanin_cevabi.append("Firma cevap vermemiştir.")
                        firmanin_cevap_verdigi_tarih.append(" ")
            print("user review:", user_review_list)

            for k in range(len(user_name)):
                """yorum puanlarının çekildiği kısım"""
                puan = 0
                for q in review_rating[k].findAll('img'):  # burada her yoruma k sayesinde tek tek gidiyor ve bu biz dizi şeklinde oluyor. q ile de her img tagı q içine aktarılıyor. Ve string değere dönüştürebiliyoruz.
                    convert = str(q)
                    if "hCCjke vzX5Ic" in convert:
                        puan = puan + 1

                details.append({"kullanici": user_name[k].text, "yorum": user_review_list[k], "yorum_puani": puan,"yorum_yapilma_tarihi": user_reviews_create_date[k].text,"firmanın cevabı":firmanin_cevabi[k],"firmanın cevap verdiği tarih":firmanin_cevap_verdigi_tarih[k]})

            print("details bilgileri:",details)
            print("details boyutu:", len(details))

            #sayfanın kaydırılması kısmı ve bilgilerin soup değişkeni içerisine atanması kısmıdır. Burada ikinci defa soup kullanıyoruz çünkü her seferinde sayfa yüklendiği için bilgiler güncelleniyor ve çekebilmek için de bizim yine html.parser  kullanmamız gerekiyor.
            ele = driver.find_element(by=By.XPATH,value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[3]')
            driver.execute_script('arguments[0].scrollBy(0, 10000);', ele) #kaydırma miktarı, diğer parametre ise sayfa içerisinde hangi alanın kaydırılacağı ile ilgili bilgiyi tutuyor.
            print("kaydırma başarılı")
            time.sleep(8) #yorumların tekrardan yüklenmesi için gereken zaman
            soup = BeautifulSoup(driver.page_source, "html.parser")
            num=num+1
        else:
            break

    for w in details:
        print(w)

    # tum_magazalar adında bir excel dosyası oluşturulması kısmı her seferinde yeniden oluşturulduğu için içindeki bilgiler tekrar etmez.
    dosya = xlsxwriter.Workbook('C:\\Users\\AwozeN\\Desktop\\deneme\\' + title + ' yorumları.xlsx')
    dosya.close()  # close ile işlemi bitiriyoruz ki dosya oluşturulsun eğer ki dosya.close() demezsek dosya oluşturulmaz.
    yeni_excel = load_workbook('C:\\Users\\AwozeN\\Desktop\\deneme\\' + title + ' yorumları.xlsx')  # okunacak olan excel dosyasının path(dosya yolu) bilgisini girdik.
    ws = yeni_excel.active

    # Başlıklar excel içerisine ekleniyor.
    ws.cell(row=1, column=1, value="Müşteri Adı")
    ws.cell(row=1, column=2, value="Yorum")
    ws.cell(row=1, column=3, value="Yorum Puanı")
    ws.cell(row=1, column=4, value="Yorum Yapılma Tarihi")
    ws.cell(row=1, column=5, value="Firmanın Cevabı")
    ws.cell(row=1, column=6, value="Firma Cevabının Tarihi")

    """ gün bilgisi şeklinde bir dizi oluşturup ekleme yaptırdım ki tarih bilgisinde karşılaştırma yaparken direk kısa şekilde bunu kullanabilelim diye"""
    gun_bilgisi=[]
    for o in range(2,13): # 2 den başlattık çünkü 1 ay önce olanları da listeye eklemek istiyoruz.
        gun_bilgisi.append(str(o)+" ay önce")
        gun_bilgisi.append(str(o)+" yıl önce")
    #bir yıl önce kısmını ayrı ekledim çünkü for içinde ekleseydim 12 tane aynı değerden olacaktı o yüzden for dışında eklettim.
    gun_bilgisi.append("bir yıl önce")
    satir=2 # 2. satırdan itibaren ekleme işlemine başlar
    for ekleme in details:
        if ekleme["yorum_yapilma_tarihi"] in gun_bilgisi:
            pass
            break
        else:
            ws.cell(row=satir,column=1,value=ekleme["kullanici"])
            ws.cell(row=satir, column=2, value=ekleme["yorum"])
            ws.cell(row=satir, column=3, value=ekleme["yorum_puani"])
            ws.cell(row=satir, column=4, value=ekleme["yorum_yapilma_tarihi"])
            ws.cell(row=satir, column=5, value=ekleme["firmanın cevabı"])
            ws.cell(row=satir, column=6, value=ekleme["firmanın cevap verdiği tarih"])
            satir=satir+1 # burada 1 arttırıyoruz ki tekrar eklemek istediğimizde bir alt satıra geçip yazsın diyedir.
    yeni_excel.save('C:\\Users\\AwozeN\\Desktop\\deneme\\' + title + ' yorumları.xlsx')



konum_linkleri=["https://goo.gl/maps/vX6doyy1CVwYoHQXA","https://goo.gl/maps/hUdjvtBtoXcLrits7","https://goo.gl/maps/XNAQ7WVY8nUmucs47"]

google_maps_yorum_bilgileri_cekme("https://goo.gl/maps/vX6doyy1CVwYoHQXA")
# for link in konum_linkleri:
#     google_maps_yorum_bilgileri_cekme(link)